import os
import shutil
from xml.dom import minidom
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd

# Caminhos
caminho_pasta = "C:\\Importador_CTE\\LER_XML"
caminho_pasta_lidos = "C:\\Importador_CTE\\LIDOS_XML"
caminho_excel = "C:\\Importador_CTE\\arquivos_CTEs_lidos2.xlsx"
caminho_txt = 'C:\\Importador_CTE\\Insert_Z12.txt'

# Listar arquivos XML na pasta
arquivos = [f for f in os.listdir(caminho_pasta) if f.endswith('.xml')]
print("Arquivos encontrados:", arquivos)

# Abrir ou criar planilha
if os.path.exists(caminho_excel):
    workbook = load_workbook(caminho_excel)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Z12_NCTE', 'Z12_SERIE', 'Z12_DATA', 'Z12_CHAVE', 'Z12_CHVNCTE', 'Z12_CGCFOR', 'NOME_TRANSPORTADORA', 'Z12_VALORT', 'Z12_CGCUPP', 'Z12_CPF', 'NOME_DESTINATARIO', 'Nome_Arquivo'])

arquivos_processados = []
arquivos_com_erro = []
arquivos_nao_gravados = []

def processar_arquivo(arquivo):
    caminho_arquivo = os.path.join(caminho_pasta, arquivo)
    resultado = []

    try:
        print(f"Processando arquivo: {arquivo}")
        with open(caminho_arquivo, 'r', encoding='utf-8') as f:
            xml = minidom.parse(f)

            # Extração de tags
            nct_tags = xml.getElementsByTagName("nCT")  # Z12_NCTE
            serie_tags = xml.getElementsByTagName("serie")  # Z12_SERIE
            infDoc_tags = xml.getElementsByTagName("infDoc")
            infProt_tags = xml.getElementsByTagName("infProt")
            dest_tags = xml.getElementsByTagName("dest")
            emit_tags = xml.getElementsByTagName("emit")
            dhEmi_tags = xml.getElementsByTagName("dhEmi")  # Z12_DATA
            vTPrest_tags = xml.getElementsByTagName("vTPrest")  # Z12_VALORT

            if not nct_tags or not nct_tags[0].firstChild:
                print(f"nCTE não encontrado em {arquivo}, pulando gravação.")
                arquivos_nao_gravados.append((arquivo, "nCTE não encontrado"))
                return []

            tags = {
                "nCT": nct_tags[0].firstChild.data.strip(),
                "serie": serie_tags[0].firstChild.data.strip() if serie_tags and serie_tags[0].firstChild else 'NAO POSSUI',
                "dhEmi": dhEmi_tags[0].firstChild.data.strip()[:10] if dhEmi_tags and dhEmi_tags[0].firstChild else 'NAO POSSUI',
                "vTPrest": vTPrest_tags[0].firstChild.data.strip().replace('.', ',') if vTPrest_tags and vTPrest_tags[0].firstChild else 'NAO POSSUI',
                "chCTe": 'NAO POSSUI'
            }

            if infProt_tags and infProt_tags[0].getElementsByTagName("chCTe"):
                chCTe_tag = infProt_tags[0].getElementsByTagName("chCTe")[0]
                if chCTe_tag and chCTe_tag.firstChild:
                    tags["chCTe"] = chCTe_tag.firstChild.data.strip()

            if dest_tags:
                cnpj_dest_value = dest_tags[0].getElementsByTagName("CNPJ")[0].firstChild.data.strip() if dest_tags[0].getElementsByTagName("CNPJ") and dest_tags[0].getElementsByTagName("CNPJ")[0].firstChild else 'NAO POSSUI'
                cpf_dest_value = dest_tags[0].getElementsByTagName("CPF")[0].firstChild.data.strip() if dest_tags[0].getElementsByTagName("CPF") and dest_tags[0].getElementsByTagName("CPF")[0].firstChild else 'NAO POSSUI'
                xnome_dest_value = dest_tags[0].getElementsByTagName("xNome")[0].firstChild.data.strip() if dest_tags[0].getElementsByTagName("xNome") and dest_tags[0].getElementsByTagName("xNome")[0].firstChild else 'NAO POSSUI'
            else:
                cnpj_dest_value = 'NAO POSSUI'
                cpf_dest_value = 'NAO POSSUI'
                xnome_dest_value = 'NAO POSSUI'

            cnpj_emit_value = emit_tags[0].getElementsByTagName("CNPJ")[0].firstChild.data.strip() if emit_tags and emit_tags[0].getElementsByTagName("CNPJ") and emit_tags[0].getElementsByTagName("CNPJ")[0].firstChild else 'NAO POSSUI'
            xnome_emit_value = emit_tags[0].getElementsByTagName("xNome")[0].firstChild.data.strip() if emit_tags and emit_tags[0].getElementsByTagName("xNome") and emit_tags[0].getElementsByTagName("xNome")[0].firstChild else 'NAO POSSUI'

            encontrou_chave = False
            if infDoc_tags:
                for infDoc_tag in infDoc_tags:
                    infNFe_tags = infDoc_tag.getElementsByTagName("infNFe")
                    for infNFe_tag in infNFe_tags:
                        chave_nFe_tags = infNFe_tag.getElementsByTagName("chave")
                        if chave_nFe_tags:
                            for chave_nFe_tag in chave_nFe_tags:
                                encontrou_chave = True
                                chave_nFe_value = chave_nFe_tag.firstChild.data.strip() if chave_nFe_tag.firstChild else 'NAO POSSUI'
                                resultado.append([tags["nCT"], tags["serie"], tags["dhEmi"], f"'{chave_nFe_value}", f"'{tags['chCTe']}", cnpj_emit_value, xnome_emit_value, tags["vTPrest"],
                                                  cnpj_dest_value, cpf_dest_value, xnome_dest_value, arquivo])

            if not encontrou_chave:
                chave_nFe_value = 'NAO POSSUI'
                resultado.append([tags["nCT"], tags["serie"], tags["dhEmi"], f"'{chave_nFe_value}", f"'{tags['chCTe']}", cnpj_emit_value, xnome_emit_value, tags["vTPrest"],
                                  cnpj_dest_value, cpf_dest_value, xnome_dest_value, arquivo])

        arquivos_processados.append(arquivo)

        return resultado

    except Exception as e:
        print(f"Erro no arquivo {arquivo}: {e}")
        arquivos_com_erro.append((arquivo, str(e)))
        return [[f'Erro ao processar o arquivo: {str(e)}', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 
                 'NAO POSSUI', 'NAO POSSUI', arquivo]]

# Processamento paralelo
with ThreadPoolExecutor(max_workers=4) as executor:
    futures = {executor.submit(processar_arquivo, arquivo): arquivo for arquivo in arquivos}
    for future in as_completed(futures):
        resultados = future.result()
        for resultado in resultados:
            if resultado:
                sheet.append(resultado)

# Ajustar largura das colunas
for column in sheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except Exception as e:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column_letter].width = adjusted_width

# Salvar o arquivo Excel
workbook.save(caminho_excel)
print("Processamento concluído. Dados salvos em:", caminho_excel)

# Executar segundo código
# Formatador e ajuste de colunas
def ajustar_largura_coluna(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

def formatar_com_zeros(valor, tamanho):
    return str(valor).zfill(tamanho)

def remover_hifens(data):
    return data.replace("-", "")

def substituir_virgula(valor):
    return valor.replace(",", ".")

def substituir_np(valor):
    return valor.replace("NAO POSSUI", "")

# Reabrir a planilha para ajustes
wb = load_workbook(caminho_excel)
ws = wb.active

ultima_linha = ws.max_row
for i in range(2, ultima_linha + 1):
    if ws.cell(row=i, column=1).value:
        ws.cell(row=i, column=14).value = "INSERT INTO Z12010 ( Z12_NCTE,Z12_SERIE,Z12_CHAVE,Z12_CHVCTE,Z12_CGCUPP,Z12_CGCFOR,Z12_DATA,Z12_VALORT,Z12_CPF,R_E_C_N_O_) VALUES ("
        valorA = formatar_com_zeros(ws.cell(row=i, column=1).value, 9)
        ws.cell(row=i, column=15).value = f"'{valorA}',"
        ws.cell(row=i, column=16).value = f"'{ws.cell(row=i, column=2).value}',"
        valorD_formatado = substituir_np(ws.cell(row=i, column=4).value)
        ws.cell(row=i, column=17).value = f"{valorD_formatado}',"
        ws.cell(row=i, column=18).value = f"{ws.cell(row=i, column=5).value}',"
        ValorH_formatado = substituir_np(ws.cell(row=i, column=9).value)
        ws.cell(row=i, column=19).value = f"'{ValorH_formatado}',"
        ws.cell(row=i, column=20).value = f"'{ws.cell(row=i, column=6).value}',"
        valorC_formatado = remover_hifens(ws.cell(row=i, column=3).value)
        ws.cell(row=i, column=21).value = f"'{valorC_formatado}',"
        valorG_formatado = substituir_virgula(ws.cell(row=i, column=8).value)
        ws.cell(row=i, column=22).value = f"'{valorG_formatado}',"
        ValorI_formatado = substituir_np(ws.cell(row=i, column=10).value)
        ws.cell(row=i, column=23).value = f"'{ValorI_formatado}',"
        ws.cell(row=i, column=24).value = "(SELECT MAX(R_E_C_N_O_)+1 FROM Z12010));"

# Ajustar largura das colunas
ajustar_largura_coluna(ws)
# Salvar arquivo Excel final
wb.save(caminho_excel)

# Remover todos os arquivos da pasta, independentemente de terem sido processados    #---------> adicionado 21/10/2024 <---------#
for arquivo in os.listdir(caminho_pasta):
    arquivo_path = os.path.join(caminho_pasta, arquivo)
    try:
        os.remove(arquivo_path)
        print(f"Arquivo {arquivo} removido com sucesso.")
    except Exception as e:
        print(f"Erro ao remover o arquivo {arquivo}: {e}")

time.sleep(0.5)

# Abrir o arquivo Excel
df = pd.read_excel(caminho_excel, usecols='N:X', skiprows=1)

# Salvar o conteúdo em um arquivo .txt
with open(caminho_txt, 'w', encoding='utf-8') as f:
    for index, row in df.iterrows():
        # Juntar os valores da linha em uma string
        linha = '\t'.join(map(str, row.values))
        f.write(linha + '\n')
        
print("Arquivo .txt gerado com sucesso.")

print("Operação concluída.")
