import os
import shutil
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time
from concurrent.futures import ThreadPoolExecutor, as_completed #---------> adicionado 16/09/2024 <---------#

# Caminho da pasta onde estão os arquivos XML
caminho_pasta = "C:\\Importador_CTE\\LER_XML"
caminho_pasta_lidos = "C:\\Importador_CTE\\LIDOS_XML"

# Caminho do arquivo Excel de saída
caminho_excel = "C:\\Importador_CTE\\arquivos_CTEs_lidos.xlsx"

# Listar todos os arquivos na pasta
arquivos = [f for f in os.listdir(caminho_pasta) if f.endswith('.xml')]
print("Arquivos encontrados:", arquivos)

# Criar um novo arquivo Excel
workbook = Workbook()
sheet = workbook.active

# Escrever o cabeçalho no arquivo Excel 
sheet.append(['Z12_NCTE', 'Z12_SERIE', 'Z12_DATA', 'Z12_CHAVE', 'Z12_CHVNCTE', 'Z12_CGCFOR', 'NOME_TRANSPORTADORA', 'Z12_VALORT', 'Z12_CGCUPP', 'Z12_CPF', 'NOME_DESTINATARIO', 'Nome_Arquivo'])

arquivos_processados = []
arquivos_com_erro = []
arquivos_nao_gravados = []                                      #---------> adicionado 21/09/2024 <---------#

def processar_arquivo(arquivo):
    caminho_arquivo = os.path.join(caminho_pasta, arquivo)
    resultado = []

    try:
        print(f"Processando arquivo: {arquivo}")
        with open(caminho_arquivo, 'r', encoding='utf-8') as f:
            xml = minidom.parse(f)

            # Extração de tags
            nct_tags = xml.getElementsByTagName("nCT")  # Z12_NCTE
            serie_tags = xml.getElementsByTagName("serie")  # Z12_SERIE             #---------> adicionado 09/09/2024 <---------#            
            infDoc_tags = xml.getElementsByTagName("infDoc")
            infProt_tags = xml.getElementsByTagName("infProt")                      #---------> adicionado 13/09/2024 <---------#                      
            dest_tags = xml.getElementsByTagName("dest")
            emit_tags = xml.getElementsByTagName("emit")
            dhEmi_tags = xml.getElementsByTagName("dhEmi")  # Z12_DATA
            vTPrest_tags = xml.getElementsByTagName("vTPrest")  # Z12_VALORT

            # Verifica se nCT foi encontrado                                        #---------> adicionado 21/09/2024 <---------#
            if not nct_tags or not nct_tags[0].firstChild:
                print(f"nCTE não encontrado em {arquivo}, pulando gravação.")
                arquivos_nao_gravados.append((arquivo, "nCTE não encontrado"))
                return []  # Retorna uma lista vazia para não gravar no Excel

            # Armazenar valores em um dicionário para melhorar o desempenho         #---------> adicionado 16/09/2024 <---------#
            tags = {
                "nCT": nct_tags[0].firstChild.data.strip(),
                "serie": serie_tags[0].firstChild.data.strip() if serie_tags and serie_tags[0].firstChild else 'NAO POSSUI',
                "dhEmi": dhEmi_tags[0].firstChild.data.strip()[:10] if dhEmi_tags and dhEmi_tags[0].firstChild else 'NAO POSSUI',
                "vTPrest": vTPrest_tags[0].firstChild.data.strip().replace('.', ',') if vTPrest_tags and vTPrest_tags[0].firstChild else 'NAO POSSUI',
                "chCTe": 'NAO POSSUI'
            }

            # Extração da tag <chCTe> dentro de <infProt>       Z12_CHVCTE            #---------> adicionado 13/09/2024 <---------#
            if infProt_tags and infProt_tags[0].getElementsByTagName("chCTe"):
                chCTe_tag = infProt_tags[0].getElementsByTagName("chCTe")[0]
                if chCTe_tag and chCTe_tag.firstChild:
                    tags["chCTe"] = chCTe_tag.firstChild.data.strip()

            # Extração de CNPJ ou CPF e xNome de dentro da tag <dest>
            if dest_tags:
                cnpj_dest_value = dest_tags[0].getElementsByTagName("CNPJ")[0].firstChild.data.strip() if dest_tags[0].getElementsByTagName("CNPJ") and dest_tags[0].getElementsByTagName("CNPJ")[0].firstChild else 'NAO POSSUI'
                cpf_dest_value = dest_tags[0].getElementsByTagName("CPF")[0].firstChild.data.strip() if dest_tags[0].getElementsByTagName("CPF") and dest_tags[0].getElementsByTagName("CPF")[0].firstChild else 'NAO POSSUI'
                xnome_dest_value = dest_tags[0].getElementsByTagName("xNome")[0].firstChild.data.strip() if dest_tags[0].getElementsByTagName("xNome") and dest_tags[0].getElementsByTagName("xNome")[0].firstChild else 'NAO POSSUI'
            else:
                cnpj_dest_value = 'NAO POSSUI'
                cpf_dest_value = 'NAO POSSUI'
                xnome_dest_value = 'NAO POSSUI'

            # Extração de CNPJ e xNome de dentro da tag <emit>
            cnpj_emit_value = emit_tags[0].getElementsByTagName("CNPJ")[0].firstChild.data.strip() if emit_tags and emit_tags[0].getElementsByTagName("CNPJ") and emit_tags[0].getElementsByTagName("CNPJ")[0].firstChild else 'NAO POSSUI'
            xnome_emit_value = emit_tags[0].getElementsByTagName("xNome")[0].firstChild.data.strip() if emit_tags and emit_tags[0].getElementsByTagName("xNome") and emit_tags[0].getElementsByTagName("xNome")[0].firstChild else 'NAO POSSUI'

            # Adicionar os valores na lista de resultados para cada chave encontrada  Z12_CHAVE      #---------> adicionado 16/09/2024 <---------#
            encontrou_chave = False
            if infDoc_tags:
                for infDoc_tag in infDoc_tags:
                    infNFe_tags = infDoc_tag.getElementsByTagName("infNFe")
                    
                    for infNFe_tag in infNFe_tags:
                        chave_nFe_tags = infNFe_tag.getElementsByTagName("chave")
                        
                        if chave_nFe_tags:
                            for chave_nFe_tag in chave_nFe_tags:
                                encontrou_chave = True
                                chave_nFe_value = chave_nFe_tag.firstChild.data.strip() if chave_nFe_tag.firstChild else 'NAO POSSUI'

                                # Adicionar os valores na lista para cada chave
                                resultado.append([tags["nCT"], tags["serie"], tags["dhEmi"], f"'{chave_nFe_value}", f"'{tags['chCTe']}", cnpj_emit_value, xnome_emit_value, tags["vTPrest"], 
                                                  cnpj_dest_value, cpf_dest_value, xnome_dest_value, arquivo])

            # Se não encontrou nenhuma chave, registrar "NAO POSSUI"                        #---------> adicionado 16/09/2024 <---------#
            if not encontrou_chave:
                chave_nFe_value = 'NAO POSSUI'
                resultado.append([tags["nCT"], tags["serie"], tags["dhEmi"], f"'{chave_nFe_value}", f"'{tags['chCTe']}", cnpj_emit_value, xnome_emit_value, tags["vTPrest"], 
                                  cnpj_dest_value, cpf_dest_value, xnome_dest_value, arquivo])

        arquivos_processados.append(arquivo)

        return resultado

    except Exception as e:
        print(f"Erro no arquivo {arquivo}: {e}")
        arquivos_com_erro.append((arquivo, str(e)))
        return [[f'Erro ao processar o arquivo: {str(e)}', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 'NAO POSSUI', 
                 'NAO POSSUI', 'NAO POSSUI', arquivo]]


# Processar arquivos em paralelo
with ThreadPoolExecutor(max_workers=4) as executor:
    futures = {executor.submit(processar_arquivo, arquivo): arquivo for arquivo in arquivos}
    
    for future in as_completed(futures):
        resultados = future.result()
        for resultado in resultados:
            if resultado:  # Somente grava se houver resultados válidos
                sheet.append(resultado)

# Ajustar a largura das colunas automaticamente                                             #---------> adicionado 21/10/2024 <---------#
for column in sheet.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)  # Obter a letra da coluna
    
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except Exception as e:
            pass
    
    adjusted_width = (max_length + 2)  # Ajusta a largura da coluna com uma margem extra
    sheet.column_dimensions[column_letter].width = adjusted_width

# Salvar o arquivo Excel após todos os arquivos terem sido processados
workbook.save(caminho_excel)
print("Processamento concluído. Dados salvos em:", caminho_excel)

# Registrar erros em um arquivo de log                                             #---------> adicionado 21/10/2024 <---------#
log_path = "C:\\Importador_CTE\\processamento_log.txt"
with open(log_path, 'w') as log_file:
    # Registrar arquivos com erro
    for arquivo in arquivos_com_erro:
        log_file.write(f"Erro no arquivo {arquivo[0]}: {arquivo[1]}\n")

    # Registrar arquivos não gravados
    for arquivo in arquivos_nao_gravados:
        log_file.write(f"Arquivo {arquivo[0]} não gravado: {arquivo[1]}\n")

# Remover todos os arquivos da pasta, independentemente de terem sido processados    #---------> adicionado 21/10/2024 <---------#
for arquivo in os.listdir(caminho_pasta):
    arquivo_path = os.path.join(caminho_pasta, arquivo)
    try:
        os.remove(arquivo_path)
        print(f"Arquivo {arquivo} removido com sucesso.")
    except Exception as e:
        print(f"Erro ao remover o arquivo {arquivo}: {e}")
