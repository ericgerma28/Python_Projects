from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
import openpyxl
from datetime import datetime
import time
import pandas as pd

caminho_txt = 'C:\\Users\\upper\\Desktop\\Custos\\Custos.txt'
caminho_arquivo = 'C:\\Users\\upper\\Desktop\\Custos\\Custo - Upper - Copia.xlsx'

# Função para substituir texto em uma faixa de linhas (para paralelizar)
def substituir_texto_intervalo(ws, start, end):
    for linha in range(start, end + 1):
        cell = ws[f'E{linha}']
        if cell.value == "Upper - MA":
            cell.value = "'0000"
        elif cell.value == "Upper - MG":
            cell.value = "'0007"
        elif cell.value == "AdoraPet":
            cell.value = "'0101"
        elif cell.value == "MCR":
            cell.value = "'0201"
        elif cell.value == "Upper - GRU":
            cell.value = "'0009"

# Função principal para substituir texto com ThreadPoolExecutor
def processar_linha(ws, i):
    if ws.cell(row=i, column=1).value:  # Verifica se há texto na célula da coluna A
        # Obtém o valor da célula da coluna E
        valor_coluna_E = ws.cell(row=i, column=5).value
        
        # Verifica o valor da coluna E e define o texto correspondente
        if valor_coluna_E in ["'0000", "'0007", "'0009"]:
            colI = "INSERT INTO Z31010 ( Z31_FILIAL,Z31_COD,Z31_UM,Z31_DESC,Z31_MARGEM,Z31_DATA,Z31_MOEDA,R_E_C_N_O_) VALUES ("
        elif valor_coluna_E == "'0101":
            colI = "INSERT INTO Z31020 ( Z31_FILIAL,Z31_COD,Z31_UM,Z31_DESC,Z31_MARGEM,Z31_DATA,Z31_MOEDA,R_E_C_N_O_) VALUES ("
        elif valor_coluna_E == "'0201":
            colI = "INSERT INTO Z31030 ( Z31_FILIAL,Z31_COD,Z31_UM,Z31_DESC,Z31_MARGEM,Z31_DATA,Z31_MOEDA,R_E_C_N_O_) VALUES ("
        else:
            colI = None  # Caso não corresponda a nenhum dos valores, você pode definir um comportamento

        if colI:
            ws.cell(row=i, column=9).value = colI

        # Preenchimento das demais colunas (mantido o restante do código original)
        valorA = ws.cell(row=i, column=1).value
        ws.cell(row=i, column=11).value = f"'{valorA}',"

        valorF = ws.cell(row=i, column=6).value
        ws.cell(row=i, column=12).value = f"'{valorF}',"

        valorB = ws.cell(row=i, column=2).value
        ws.cell(row=i, column=13).value = f"'{valorB}',"

        valorC = ws.cell(row=i, column=3).value
        if isinstance(valorC, float):
            valorC_formatado = f"{valorC:.2f}".replace(",", ".")
        else:
            valorC_formatado = str(valorC).replace(",", ".")

        ws.cell(row=i, column=14).value = f"'{valorC_formatado}',"

        valorE = ws.cell(row=i, column=5).value
        ws.cell(row=i, column=10).value = f"{valorE}',"

        valorD = ws.cell(row=i, column=4).value
        if isinstance(valorD, datetime):
            data_formatada = valorD.strftime('%Y%m')
            ws.cell(row=i, column=15).value = f"'{data_formatada}',"

        colP = "'1',"
        ws.cell(row=i, column=16).value = colP

        if valor_coluna_E in ["'0000", "'0007", "'0009"]:
            valorQ = "(SELECT MAX(R_E_C_N_O_)+1 FROM Z31010));"
        elif valor_coluna_E == "'0101":
            valorQ = "(SELECT MAX(R_E_C_N_O_)+1 FROM Z31020));"
        elif valor_coluna_E == "'0201":
            valorQ = "(SELECT MAX(R_E_C_N_O_)+1 FROM Z31030));"
        else:
            valorQ = None  # Caso não corresponda a nenhum dos valores, você pode definir um comportamento

        if valorQ:
            ws.cell(row=i, column=17).value = valorQ


# Função para ajustar a largura das colunas
def ajustar_largura_coluna(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Pega a letra da coluna
        for cell in col:
            try: 
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Adiciona uma margem
        ws.column_dimensions[column].width = adjusted_width

# Função principal para processar custos com ThreadPoolExecutor
def custos():
    caminho_arquivo = 'C:\\Users\\upper\\Desktop\\Custos\\Custo - Upper - Copia.xlsx'
    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active
    last_row = ws.max_row
    print(f"Última linha com dados: {last_row}")

    # Usa ThreadPoolExecutor para paralelizar o trabalho por linha
    with ThreadPoolExecutor() as executor:
        futuros = []
        for i in range(2, last_row + 1):
            futuros.append(executor.submit(processar_linha, ws, i))
        
        # Aguarda todas as threads terminarem
        for futuro in futuros:
            futuro.result()

    # Ajusta a largura das colunas após o processamento
    ajustar_largura_coluna(ws)

    # Salva o arquivo após todas as modificações
    print("Salvando o arquivo...")
    wb.save(caminho_arquivo)
    print("Arquivo salvo com sucesso!")

# Substitui o texto em um intervalo de linhas da planilha
wb = openpyxl.load_workbook(caminho_arquivo)
ws = wb.active

# Executa a substituição
substituir_texto_intervalo(ws, 2, ws.max_row)

# Salva o arquivo após a substituição
wb.save(caminho_arquivo)
print("Substituição realizada e arquivo salvo.")

# Agora processa os custos
custos()

time.sleep(0.5)

# Abrir o arquivo Excel
df = pd.read_excel(caminho_arquivo, usecols='I:Q', skiprows=1)

# Salvar o conteúdo em um arquivo .txt
with open(caminho_txt, 'w', encoding='utf-8') as f:
    for index, row in df.iterrows():
        # Juntar os valores da linha em uma string
        linha = '\t'.join(map(str, row.values))
        f.write(linha + '\n')

print("Arquivo .txt gerado com sucesso.")
print("Operação concluída.")
